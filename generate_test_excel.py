#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
結合テスト項目書をExcel形式で出力するスクリプト
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def create_test_excel():
    """結合テスト項目書のExcelファイルを作成"""

    wb = Workbook()

    # シート1: テスト項目一覧
    ws_summary = wb.active
    ws_summary.title = "テスト項目一覧"

    # シート2: 画面操作テスト
    ws_api = wb.create_sheet("画面操作テスト")

    # シート3: E2Eテスト
    ws_e2e = wb.create_sheet("E2Eテスト")

    # シート4: エラーケーステスト
    ws_error = wb.create_sheet("エラーケーステスト")

    # シート5: データ整合性テスト
    ws_integrity = wb.create_sheet("データ整合性テスト")

    # シート6: テスト結果記録
    ws_results = wb.create_sheet("テスト結果記録")

    # スタイル定義
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # ===== シート1: テスト項目一覧 =====
    ws_summary.merge_cells('A1:F1')
    ws_summary['A1'] = '学習用ポモドーロアプリ 結合テスト項目一覧'
    ws_summary['A1'].font = Font(bold=True, size=16)
    ws_summary['A1'].alignment = center_align

    ws_summary['A3'] = '作成日'
    ws_summary['B3'] = datetime.now().strftime('%Y年%m月%d日')
    ws_summary['A4'] = '作成者'
    ws_summary['B4'] = '[名前]'

    headers = ['No', 'テスト項目ID', 'テスト項目名', 'テスト種別', '優先度', '備考']
    ws_summary.append([])
    ws_summary.append(headers)

    # ヘッダーのスタイル設定
    for col in range(1, len(headers) + 1):
        cell = ws_summary.cell(row=6, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    # テスト項目データ
    test_items = [
        [1, 'IT-001', 'Spotify認証URL取得（画面確認）', '画面操作', '高', ''],
        [2, 'IT-002', 'Spotify認証コールバック処理（画面確認）', '画面操作', '高', ''],
        [3, 'IT-003', 'プレイリスト一覧取得（画面確認）', '画面操作', '高', ''],
        [4, 'IT-004', 'プレイリストトラック取得（画面確認）', '画面操作', '中', ''],
        [5, 'IT-005', '設定情報取得（画面確認）', '画面操作', '高', ''],
        [6, 'IT-006', '設定情報更新（画面確認）', '画面操作', '高', ''],
        [7, 'IT-007', 'セッション作成（画面確認）', '画面操作', '高', ''],
        [8, 'IT-008', 'セッション更新（画面確認）', '画面操作', '高', ''],
        [9, 'IT-009', 'セッション一覧取得（画面確認）', '画面操作', '中', ''],
        [10, 'IT-010', 'Spotify認証フロー（E2E）', 'E2E', '高', ''],
        [11, 'IT-011', 'タイマー開始・終了フロー（E2E）', 'E2E', '高', ''],
        [12, 'IT-012', '設定変更フロー（E2E）', 'E2E', '高', ''],
        [13, 'IT-013', 'プレイリスト選択フロー（E2E）', 'E2E', '中', ''],
        [14, 'IT-014', 'バリデーションエラーテスト', 'エラー', '高', ''],
        [15, 'IT-015', '認証エラーテスト', 'エラー', '高', ''],
        [16, 'IT-016', 'データ整合性テスト', '整合性', '中', ''],
    ]

    for row_data in test_items:
        ws_summary.append(row_data)
        for col in range(1, len(row_data) + 1):
            cell = ws_summary.cell(row=ws_summary.max_row, column=col)
            cell.border = border
            if col == 1:
                cell.alignment = center_align
            else:
                cell.alignment = left_align

    # 列幅調整
    ws_summary.column_dimensions['A'].width = 8
    ws_summary.column_dimensions['B'].width = 15
    ws_summary.column_dimensions['C'].width = 40
    ws_summary.column_dimensions['D'].width = 15
    ws_summary.column_dimensions['E'].width = 10
    ws_summary.column_dimensions['F'].width = 30

    # ===== シート2: 画面操作テスト詳細 =====
    api_tests = [
        {
            'id': 'IT-001',
            'name': 'Spotify認証URL取得（画面確認）',
            'endpoint': 'GET /api/spotify/auth-url',
            'purpose': '画面操作でSpotify認証URLが正しく取得されることを確認',
            'steps': '1. ブラウザでhttp://localhost:5173にアクセス\n2. 開発者ツール（F12）を開き、Networkタブを表示\n3. 「Spotifyでログイン」ボタンをクリック\n4. Networkタブで/api/spotify/auth-urlのリクエストを確認\n5. 認証ページにリダイレクトされることを確認',
            'expected': '【画面】「Spotifyでログイン」ボタンをクリックすると認証ページに遷移する\n【Network】GET /api/spotify/auth-urlが送信される、ステータスコード: 200 OK',
            'request': 'なし',
            'response': '認証ページにリダイレクトされる'
        },
        {
            'id': 'IT-002',
            'name': 'Spotify認証コールバック処理（画面確認）',
            'endpoint': 'POST /api/spotify/callback',
            'purpose': '認証後のコールバック処理が正常に完了し、トークンが保存されることを確認',
            'steps': '1. 認証ページで認証を完了（モックモード時は自動的にコールバックURLに戻る）\n2. 開発者ツール（F12）のNetworkタブを確認\n3. /api/spotify/callbackまたは/api/spotify/tokenのリクエストを確認\n4. ApplicationタブでLocalStorageを確認',
            'expected': '【画面】コールバックURLに戻り、エラーが表示されない\n【Network】POST /api/spotify/callbackが送信される、ステータスコード: 200 OK\n【LocalStorage】spotifyAccessTokenが保存されている',
            'request': '認証コード（自動）',
            'response': '認証状態が「認証済み」と表示される'
        },
        {
            'id': 'IT-003',
            'name': 'プレイリスト一覧取得（画面確認）',
            'endpoint': 'GET /api/spotify/playlists',
            'purpose': '画面操作でプレイリスト一覧が正しく表示されることを確認',
            'steps': '1. ブラウザでhttp://localhost:5173にアクセス\n2. 開発者ツール（F12）を開き、Networkタブを表示\n3. 「プレイリスト選択」ボタンをクリック\n4. プレイリスト選択モーダルが表示されることを確認\n5. Networkタブで/api/spotify/playlistsのリクエストを確認',
            'expected': '【画面】「プレイリスト選択」ボタンをクリックするとモーダルが表示される\nプレイリスト一覧がモーダルに表示される（モックモード時は3件）\n【Network】GET /api/spotify/playlistsが送信される、ステータスコード: 200 OK',
            'request': 'ヘッダー: Authorization: Bearer {accessToken}',
            'response': 'プレイリスト一覧がモーダルに表示される'
        },
        {
            'id': 'IT-004',
            'name': 'プレイリストトラック取得（画面確認）',
            'endpoint': 'GET /api/spotify/playlists/{playlistId}/tracks',
            'purpose': 'プレイリスト選択後、トラック情報が正しく表示されることを確認',
            'steps': '1. ブラウザでhttp://localhost:5173にアクセス\n2. 「プレイリスト選択」ボタンをクリック\n3. プレイリストを1つ選択\n4. 開発者ツール（F12）のNetworkタブを確認\n5. 画面にトラック情報（曲名、アーティスト名、アートワーク）が表示されることを確認',
            'expected': '【画面】プレイリストを選択すると、選択したプレイリスト名が表示される\n音楽情報カードに曲名、アーティスト名、アートワークが表示される\n【Network】GET /api/spotify/playlists/{playlistId}/tracksが送信される（実装されている場合）',
            'request': 'プレイリスト選択操作',
            'response': '音楽情報カードにトラック情報が表示される'
        },
        {
            'id': 'IT-005',
            'name': '設定情報取得（画面確認）',
            'endpoint': 'GET /api/settings/{deviceId}',
            'purpose': '設定画面で設定情報が正しく表示されることを確認',
            'steps': '1. ブラウザでhttp://localhost:5173/settingsにアクセス\n2. 開発者ツール（F12）を開き、Networkタブを表示\n3. 設定画面が表示されることを確認\n4. Networkタブで/api/settings/{deviceId}のリクエストを確認\n5. 設定値（作業時間、休憩時間）が表示されることを確認',
            'expected': '【画面】設定画面が正常に表示される\n作業時間と休憩時間の入力欄に値が表示される（デフォルト値: 作業25分、休憩5分）\n【Network】GET /api/settings/{deviceId}が送信される、ステータスコード: 200 OK',
            'request': 'なし',
            'response': '設定値が入力欄に表示される'
        },
        {
            'id': 'IT-006',
            'name': '設定情報更新（画面確認）',
            'endpoint': 'PUT /api/settings',
            'purpose': '設定画面で設定を変更し、保存できることを確認',
            'steps': '1. ブラウザでhttp://localhost:5173/settingsにアクセス\n2. 開発者ツール（F12）を開き、Networkタブを表示\n3. 作業時間を30分、休憩時間を10分に変更\n4. 「保存」ボタンをクリック\n5. NetworkタブでPUT /api/settingsのリクエストを確認\n6. タイマーページに戻り、設定が反映されていることを確認',
            'expected': '【画面】「保存」ボタンをクリックすると、エラーメッセージが表示されない\nタイマーページに戻ると、変更した設定値が反映されている\nページをリロードしても設定が保持されている\n【Network】PUT /api/settingsが送信される、ステータスコード: 200 OK',
            'request': '{"deviceId": "...", "workMinutes": 30, "breakMinutes": 10, ...}',
            'response': '設定が保存され、タイマーページに反映される'
        },
        {
            'id': 'IT-007',
            'name': 'セッション作成（画面確認）',
            'endpoint': 'POST /api/sessions',
            'purpose': 'タイマー開始時にセッションが正しく作成されることを確認',
            'steps': '1. ブラウザでhttp://localhost:5173にアクセス\n2. 開発者ツール（F12）を開き、Networkタブを表示\n3. 「開始」ボタンをクリック\n4. NetworkタブでPOST /api/sessionsのリクエストを確認\n5. タイマーが開始されることを確認',
            'expected': '【画面】「開始」ボタンをクリックするとタイマーが開始される\nタイマーの表示がカウントダウンする\nエラーが表示されない\n【Network】POST /api/sessionsが送信される、ステータスコード: 201 Created',
            'request': '{"deviceId": "...", "sessionType": "WORK", "durationMinutes": 25, ...}',
            'response': 'タイマーが開始される'
        },
        {
            'id': 'IT-008',
            'name': 'セッション更新（画面確認）',
            'endpoint': 'PUT /api/sessions/{id}',
            'purpose': 'タイマー終了時にセッションが正しく更新されることを確認',
            'steps': '1. タイマーを開始（IT-007）\n2. 開発者ツール（F12）を開き、Networkタブを表示\n3. タイマーが終了するまで待つ（テスト用に1分に設定推奨）\n4. タイマー終了後、NetworkタブでPUT /api/sessions/{id}のリクエストを確認\n5. 終了モーダルが表示されることを確認',
            'expected': '【画面】タイマーが0:00になると終了モーダルが表示される\nエラーが表示されない\n休憩時間のタイマーが自動的に開始される（実装されている場合）\n【Network】PUT /api/sessions/{id}が送信される、ステータスコード: 200 OK',
            'request': '{"deviceId": "...", "sessionType": "WORK", "completed": true}',
            'response': '終了モーダルが表示される'
        },
        {
            'id': 'IT-009',
            'name': 'セッション一覧取得（画面確認）',
            'endpoint': 'GET /api/sessions?deviceId={deviceId}',
            'purpose': 'セッション履歴が正しく表示されることを確認（実装されている場合）',
            'steps': '1. タイマーを複数回開始・終了してセッションを作成\n2. セッション履歴画面（実装されている場合）にアクセス\n3. 開発者ツール（F12）を開き、Networkタブを表示\n4. NetworkタブでGET /api/sessions?deviceId={deviceId}のリクエストを確認\n5. セッション履歴が表示されることを確認',
            'expected': '【画面】セッション履歴画面にセッション一覧が表示される\nセッションが作成日時の降順（新しい順）で表示される\n【Network】GET /api/sessions?deviceId={deviceId}が送信される、ステータスコード: 200 OK',
            'request': 'クエリパラメータ: deviceId={deviceId}',
            'response': 'セッション履歴が表示される（実装されている場合）'
        },
    ]

    create_detail_sheet(ws_api, '画面操作テスト', api_tests, header_fill, header_font, border, center_align, left_align)

    # ===== シート3: E2Eテスト詳細 =====
    e2e_tests = [
        {
            'id': 'IT-010',
            'name': 'Spotify認証フロー（E2E）',
            'purpose': 'フロントエンドからバックエンドを経由してSpotify認証が完了することを確認',
            'steps': '1. ブラウザでhttp://localhost:5173にアクセス\n2. 「Spotifyでログイン」ボタンをクリック\n3. 認証URLが取得され、リダイレクトされることを確認\n4. コールバックURLに戻り、トークンが取得されることを確認\n5. LocalStorageにトークンが保存されることを確認',
            'expected': '認証URLが正しく取得される\nコールバック処理が正常に完了する\nアクセストークンがLocalStorageに保存される'
        },
        {
            'id': 'IT-011',
            'name': 'タイマー開始・終了フロー（E2E）',
            'purpose': 'タイマー開始から終了まで、セッションが正しく記録されることを確認',
            'steps': '1. ブラウザでhttp://localhost:5173にアクセス\n2. 設定画面で作業時間を1分に設定（テスト用）\n3. タイマーページに戻り、「開始」ボタンをクリック\n4. セッション作成APIが呼ばれることを確認\n5. タイマーがカウントダウンすることを確認\n6. タイマー終了後、セッション更新APIが呼ばれることを確認',
            'expected': 'セッション作成APIがPOST /api/sessionsで呼ばれる\nタイマーが正しくカウントダウンする\nタイマー終了後、セッション更新APIがPUT /api/sessions/{id}で呼ばれる'
        },
        {
            'id': 'IT-012',
            'name': '設定変更フロー（E2E）',
            'purpose': '設定画面で設定を変更し、バックエンドに保存されることを確認',
            'steps': '1. ブラウザでhttp://localhost:5173/settingsにアクセス\n2. 作業時間を30分、休憩時間を10分に変更\n3. 「保存」ボタンをクリック\n4. 設定更新APIが呼ばれることを確認\n5. タイマーページに戻り、設定が反映されていることを確認',
            'expected': '設定更新APIがPUT /api/settingsで呼ばれる\nタイマーページで設定が反映される\nデータベースに設定が保存される'
        },
        {
            'id': 'IT-013',
            'name': 'プレイリスト選択フロー（E2E）',
            'purpose': 'Spotify認証後、プレイリストを選択できることを確認',
            'steps': '1. ブラウザでhttp://localhost:5173にアクセス\n2. Spotify認証を完了\n3. 「プレイリスト選択」ボタンをクリック\n4. プレイリスト一覧取得APIが呼ばれることを確認\n5. プレイリスト一覧が表示されることを確認\n6. プレイリストを選択',
            'expected': 'プレイリスト一覧取得APIがGET /api/spotify/playlistsで呼ばれる\nプレイリスト一覧がモーダルに表示される\nプレイリストを選択すると、選択したプレイリスト名が表示される'
        },
    ]

    create_e2e_sheet(ws_e2e, 'E2Eテスト', e2e_tests, header_fill, header_font, border, center_align, left_align)

    # ===== シート4: エラーケーステスト詳細 =====
    error_tests = [
        {
            'id': 'IT-014',
            'name': 'バリデーションエラーテスト（画面確認）',
            'purpose': '不正な入力値に対して適切なエラーメッセージが表示されることを確認',
            'steps': '1. ブラウザでhttp://localhost:5173/settingsにアクセス\n2. 開発者ツール（F12）を開き、Networkタブを表示\n3. 以下の不正な値を入力して「保存」ボタンをクリック：\n   - 作業時間を100分（上限超過）\n   - 作業時間を0分（下限未満）\n   - 休憩時間を35分（上限超過）\n4. エラーメッセージが表示されることを確認\n5. Networkタブでエラーレスポンスを確認',
            'expected': '【画面】不正な値を入力して「保存」をクリックするとエラーメッセージが表示される\nエラーメッセージに具体的な内容が含まれる\n保存が実行されない\n【Network】PUT /api/settingsが送信される、ステータスコード: 400 Bad Request'
        },
        {
            'id': 'IT-015',
            'name': '認証エラーテスト（画面確認）',
            'purpose': '認証が完了していない状態でプレイリスト選択を試みた際のエラー処理を確認',
            'steps': '1. ブラウザでhttp://localhost:5173にアクセス\n2. LocalStorageからSpotify認証情報を削除（ApplicationタブでspotifyAccessTokenを削除）\n3. 開発者ツール（F12）を開き、Networkタブを表示\n4. 「プレイリスト選択」ボタンをクリック（表示されている場合）\n5. エラーメッセージまたは認証要求が表示されることを確認\n6. Networkタブでエラーレスポンスを確認',
            'expected': '【画面】認証が完了していない場合、「Spotifyでログイン」モーダルが表示される\nまたは、エラーメッセージが表示される\nプレイリスト選択ができない状態になる\n【Network】エラーが発生する場合、ステータスコード: 401 Unauthorized または 500 Internal Server Error（モックモード時）'
        },
    ]

    create_e2e_sheet(ws_error, 'エラーケーステスト', error_tests, header_fill, header_font, border, center_align, left_align)

    # ===== シート5: データ整合性テスト詳細 =====
    integrity_tests = [
        {
            'id': 'IT-016',
            'name': 'データ整合性テスト（画面確認）',
            'purpose': '画面操作で変更したデータが正しく保持されることを確認',
            'steps': '1. 設定画面で作業時間を30分、休憩時間を10分に変更して保存\n2. ページをリロード（F5）\n3. 設定画面またはタイマーページで設定値が保持されていることを確認\n4. タイマーを開始してセッションを作成\n5. ページをリロード\n6. タイマーの状態が正しく復元されることを確認（実装されている場合）',
            'expected': '【画面】設定を変更して保存後、ページをリロードしても設定値が保持されている\nタイマーを開始後、ページをリロードしてもセッション情報が保持されている（実装されている場合）\nデータが失われていない\n【Network】ページリロード時、GET /api/settings/{deviceId}が送信される、ステータスコード: 200 OK'
        },
    ]

    create_e2e_sheet(ws_integrity, 'データ整合性テスト', integrity_tests, header_fill, header_font, border, center_align, left_align)

    # ===== シート6: テスト結果記録 =====
    ws_results.merge_cells('A1:F1')
    ws_results['A1'] = 'テスト実行記録表'
    ws_results['A1'].font = title_font
    ws_results['A1'].alignment = center_align

    result_headers = ['テスト項目ID', '実行日時', '実行者', '結果', '不具合ID', '備考']
    ws_results.append([])
    ws_results.append(result_headers)

    for col in range(1, len(result_headers) + 1):
        cell = ws_results.cell(row=3, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    # テスト項目IDを追加
    for item in test_items:
        row = [item[1], '', '', '', '', '']
        ws_results.append(row)
        for col in range(1, len(row) + 1):
            cell = ws_results.cell(row=ws_results.max_row, column=col)
            cell.border = border
            if col == 4:  # 結果列
                cell.alignment = center_align
            else:
                cell.alignment = left_align

    # 列幅調整
    ws_results.column_dimensions['A'].width = 15
    ws_results.column_dimensions['B'].width = 20
    ws_results.column_dimensions['C'].width = 15
    ws_results.column_dimensions['D'].width = 10
    ws_results.column_dimensions['E'].width = 15
    ws_results.column_dimensions['F'].width = 40

    # テストサマリー
    ws_results.append([])
    ws_results.append(['テストサマリー'])
    ws_results.cell(row=ws_results.max_row, column=1).font = title_font

    summary_headers = ['項目', '件数']
    ws_results.append(summary_headers)
    for col in range(1, len(summary_headers) + 1):
        cell = ws_results.cell(row=ws_results.max_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    summary_data = [
        ['総テスト項目数', len(test_items)],
        ['合格件数', ''],
        ['不合格件数', ''],
        ['未実施件数', ''],
        ['合格率', ''],
    ]

    for row_data in summary_data:
        ws_results.append(row_data)
        for col in range(1, len(row_data) + 1):
            cell = ws_results.cell(row=ws_results.max_row, column=col)
            cell.border = border
            cell.alignment = left_align if col == 1 else center_align

    # ファイル保存
    filename = '結合テスト項目書_学習用ポモドーロアプリ.xlsx'
    wb.save(filename)
    print(f'Excelファイルを作成しました: {filename}')
    return filename

def create_detail_sheet(ws, title, tests, header_fill, header_font, border, center_align, left_align):
    """詳細テストシートを作成"""
    ws.merge_cells('A1:H1')
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center_align

    headers = ['テスト項目ID', 'テスト項目名', 'エンドポイント', 'テスト目的', 'テスト手順', '期待結果', 'リクエスト例', 'レスポンス例']
    ws.append([])
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    for test in tests:
        row = [
            test['id'],
            test['name'],
            test.get('endpoint', ''),
            test['purpose'],
            test['steps'],
            test['expected'],
            test.get('request', ''),
            test.get('response', '')
        ]
        ws.append(row)
        for col in range(1, len(row) + 1):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.border = border
            cell.alignment = left_align

    # 列幅調整
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 40
    ws.column_dimensions['H'].width = 40

def create_e2e_sheet(ws, title, tests, header_fill, header_font, border, center_align, left_align):
    """E2Eテストシートを作成"""
    ws.merge_cells('A1:E1')
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center_align

    headers = ['テスト項目ID', 'テスト項目名', 'テスト目的', 'テスト手順', '期待結果']
    ws.append([])
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    for test in tests:
        row = [
            test['id'],
            test['name'],
            test['purpose'],
            test['steps'],
            test['expected']
        ]
        ws.append(row)
        for col in range(1, len(row) + 1):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.border = border
            cell.alignment = left_align

    # 列幅調整
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 50

if __name__ == '__main__':
    try:
        create_test_excel()
    except ImportError:
        print('エラー: openpyxlライブラリが必要です。')
        print('インストール方法: pip install openpyxl')
    except Exception as e:
        print(f'エラーが発生しました: {e}')

