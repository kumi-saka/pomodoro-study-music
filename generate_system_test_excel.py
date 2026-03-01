#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
総合テスト項目書をExcel形式で出力するスクリプト（ユーザーシナリオベース）
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def create_system_test_excel():
    """総合テスト項目書のExcelファイルを作成"""

    wb = Workbook()

    # シート1: テスト項目一覧
    ws_summary = wb.active
    ws_summary.title = "テスト項目一覧"

    # シート2: ユーザーシナリオ
    ws_scenarios = wb.create_sheet("ユーザーシナリオ")

    # シート3: シナリオテスト詳細
    ws_scenario_tests = wb.create_sheet("シナリオテスト詳細")

    # シート4: ユーザビリティテスト
    ws_usability = wb.create_sheet("ユーザビリティテスト")

    # シート5: システム全体動作確認
    ws_system = wb.create_sheet("システム全体動作確認")

    # シート6: テスト結果記録
    ws_results = wb.create_sheet("テスト結果記録")

    # スタイル定義
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14)
    scenario_fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # ===== シート1: テスト項目一覧 =====
    ws_summary.merge_cells('A1:G1')
    ws_summary['A1'] = '学習用ポモドーロアプリ 総合テスト項目一覧'
    ws_summary['A1'].font = Font(bold=True, size=16)
    ws_summary['A1'].alignment = center_align

    ws_summary['A3'] = '作成日'
    ws_summary['B3'] = datetime.now().strftime('%Y年%m月%d日')
    ws_summary['A4'] = '作成者'
    ws_summary['B4'] = '[名前]'

    headers = ['No', 'テスト項目ID', 'テスト項目名', 'シナリオID', 'テスト種別', '優先度', '所要時間（目安）']
    ws_summary.append([])
    ws_summary.append(headers)

    # ヘッダーのスタイル設定
    for col in range(1, len(headers) + 1):
        cell = ws_summary.cell(row=6, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    # テスト項目データ（ユーザーシナリオベース）
    test_items = [
        [1, 'ST-001', '初回利用フロー', 'SCENARIO-001', 'シナリオ', '高', '10分'],
        [2, 'ST-002', 'Spotify認証から音楽再生までの完全フロー', 'SCENARIO-002', 'シナリオ', '高', '15分'],
        [3, 'ST-003', '設定変更からタイマー実行までのフロー', 'SCENARIO-003', 'シナリオ', '高', '10分'],
        [4, 'ST-004', '複数セッション実行フロー', 'SCENARIO-004', 'シナリオ', '高', '15分'],
        [5, 'ST-005', '作業→休憩→作業の連続実行フロー', 'SCENARIO-004', 'シナリオ', '高', '20分'],
        [6, 'ST-006', 'ブラウザリロード時の状態復元', 'SCENARIO-001,003', 'シナリオ', '中', '10分'],
        [7, 'ST-007', 'エラー発生時の動作確認', 'SCENARIO-005', 'シナリオ', '高', '15分'],
        [8, 'ST-008', 'モバイル対応確認', '-', 'ユーザビリティ', '中', '10分'],
        [9, 'ST-009', '画面遷移の確認', '-', 'ユーザビリティ', '中', '5分'],
        [10, 'ST-010', 'データ整合性の確認（複数操作後）', '-', 'システム全体', '高', '15分'],
    ]

    for row_data in test_items:
        ws_summary.append(row_data)
        for col in range(1, len(row_data) + 1):
            cell = ws_summary.cell(row=ws_summary.max_row, column=col)
            cell.border = border
            if col == 1 or col == 5 or col == 6:
                cell.alignment = center_align
            else:
                cell.alignment = left_align

    # 列幅調整
    ws_summary.column_dimensions['A'].width = 8
    ws_summary.column_dimensions['B'].width = 12
    ws_summary.column_dimensions['C'].width = 45
    ws_summary.column_dimensions['D'].width = 20
    ws_summary.column_dimensions['E'].width = 15
    ws_summary.column_dimensions['F'].width = 10
    ws_summary.column_dimensions['G'].width = 15

    # ===== シート2: ユーザーシナリオ =====
    ws_scenarios.merge_cells('A1:D1')
    ws_scenarios['A1'] = 'ユーザーシナリオ一覧'
    ws_scenarios['A1'].font = title_font
    ws_scenarios['A1'].alignment = center_align

    scenario_headers = ['シナリオID', 'シナリオ名', '概要', '関連テスト項目ID']
    ws_scenarios.append([])
    ws_scenarios.append(scenario_headers)

    for col in range(1, len(scenario_headers) + 1):
        cell = ws_scenarios.cell(row=3, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    scenarios = [
        ['SCENARIO-001', '初回利用ユーザー', 'アプリを初めて使用するユーザーの流れ。デバイスID生成、デフォルト設定取得、タイマー実行まで。', 'ST-001, ST-006'],
        ['SCENARIO-002', '通常利用ユーザー（Spotify認証済み）', 'Spotify認証からプレイリスト選択、音楽再生まで。音楽を聴きながら学習するユーザー。', 'ST-002'],
        ['SCENARIO-003', '設定カスタマイズユーザー', '自分の作業スタイルに合わせてタイマー時間をカスタマイズするユーザー。', 'ST-003, ST-006'],
        ['SCENARIO-004', '長時間学習ユーザー', '複数のポモドーロセッションを連続して実行するユーザー。作業→休憩→作業の繰り返し。', 'ST-004, ST-005'],
        ['SCENARIO-005', 'エラー発生時のユーザー', 'エラーが発生した際の動作確認。バリデーションエラー、ネットワークエラーなど。', 'ST-007'],
    ]

    for row_data in scenarios:
        ws_scenarios.append(row_data)
        for col in range(1, len(row_data) + 1):
            cell = ws_scenarios.cell(row=ws_scenarios.max_row, column=col)
            cell.border = border
            cell.alignment = left_align
            if col == 1:
                cell.fill = scenario_fill

    # 列幅調整
    ws_scenarios.column_dimensions['A'].width = 18
    ws_scenarios.column_dimensions['B'].width = 30
    ws_scenarios.column_dimensions['C'].width = 60
    ws_scenarios.column_dimensions['D'].width = 25

    # ===== シート3: シナリオテスト詳細 =====
    scenario_tests = [
        {
            'id': 'ST-001',
            'scenario': 'SCENARIO-001',
            'name': '初回利用フロー',
            'purpose': '初回利用者がアプリを使用開始してから、タイマーを実行できるまでの一連の流れを確認',
            'steps': '1. ブラウザでhttp://localhost:5173にアクセス\n2. 開発者ツール（F12）のApplicationタブでLocalStorageを確認\n3. タイマーページが表示されることを確認\n4. 設定画面（/settings）にアクセス\n5. 設定値がデフォルト値（作業25分、休憩5分）で表示されることを確認\n6. タイマーページに戻る\n7. 「開始」ボタンをクリック\n8. タイマーが開始されることを確認\n9. NetworkタブでAPI呼び出しを確認',
            'expected_screen': '初回アクセス時にエラーが表示されない\nLocalStorageにdeviceIdが自動生成される\n設定画面でデフォルト値が表示される\nタイマーが正常に開始される\nタイマーがカウントダウンする',
            'expected_api': 'GET /api/settings/{deviceId}が呼ばれる（設定取得）\nPOST /api/sessionsが呼ばれる（セッション作成）\nすべてのAPIが正常にレスポンスを返す（ステータスコード: 200/201）'
        },
        {
            'id': 'ST-002',
            'scenario': 'SCENARIO-002',
            'name': 'Spotify認証から音楽再生までの完全フロー',
            'purpose': 'Spotify認証からプレイリスト選択、タイマー開始、音楽再生までの一連の流れを確認',
            'steps': '1. ブラウザでhttp://localhost:5173にアクセス\n2. 「Spotifyでログイン」ボタンをクリック\n3. 認証ページに遷移することを確認（モックモード時は自動的にコールバックURLに戻る）\n4. コールバックURLに戻り、認証が完了することを確認\n5. 「プレイリスト選択」ボタンが表示されることを確認\n6. 「プレイリスト選択」ボタンをクリック\n7. プレイリスト一覧がモーダルに表示されることを確認\n8. プレイリストを1つ選択\n9. 選択したプレイリスト名が表示されることを確認\n10. 音楽情報カードに曲名、アーティスト名、アートワークが表示されることを確認\n11. 「開始」ボタンをクリック\n12. タイマーが開始され、音楽が再生されることを確認（モックモード時は再生UIが表示される）',
            'expected_screen': 'Spotify認証が正常に完了する\nプレイリスト一覧が表示される\nプレイリストを選択できる\n音楽情報が表示される\nタイマー開始時に音楽再生が開始される（または再生UIが表示される）',
            'expected_api': 'GET /api/spotify/auth-urlが呼ばれる\nPOST /api/spotify/callbackが呼ばれる\nGET /api/spotify/playlistsが呼ばれる\nPOST /api/sessionsが呼ばれる\nすべてのAPIが正常にレスポンスを返す'
        },
        {
            'id': 'ST-003',
            'scenario': 'SCENARIO-003',
            'name': '設定変更からタイマー実行までのフロー',
            'purpose': '設定を変更してから、変更した設定でタイマーを実行できることを確認',
            'steps': '1. ブラウザでhttp://localhost:5173/settingsにアクセス\n2. 作業時間を30分、休憩時間を10分に変更\n3. 「保存」ボタンをクリック\n4. 保存成功のメッセージが表示されることを確認（実装されている場合）\n5. タイマーページ（/）に戻る\n6. タイマーの初期表示が30:00になっていることを確認\n7. 「開始」ボタンをクリック\n8. タイマーが30分からカウントダウンすることを確認\n9. タイマーが終了するまで待つ（テスト用に1分に設定して再テスト推奨）\n10. タイマー終了後、休憩タイマーが10分で開始されることを確認',
            'expected_screen': '設定が正常に保存される\nタイマーページで変更した設定が反映される\nタイマーが変更した時間で実行される\n休憩タイマーも変更した時間で実行される',
            'expected_api': 'PUT /api/settingsが呼ばれる\nGET /api/settings/{deviceId}が呼ばれる（タイマーページ表示時）\nPOST /api/sessionsが呼ばれる（タイマー開始時）\nPUT /api/sessions/{id}が呼ばれる（タイマー終了時）'
        },
        {
            'id': 'ST-004',
            'scenario': 'SCENARIO-004',
            'name': '複数セッション実行フロー',
            'purpose': '複数のセッションを連続して実行し、それぞれが正しく記録されることを確認',
            'steps': '1. ブラウザでhttp://localhost:5173にアクセス\n2. 開発者ツール（F12）のNetworkタブを開く\n3. 「開始」ボタンをクリックしてセッション1を開始\n4. タイマーが終了するまで待つ\n5. 終了モーダルが表示されることを確認\n6. 「開始」ボタンをクリックしてセッション2を開始\n7. タイマーが終了するまで待つ\n8. 「開始」ボタンをクリックしてセッション3を開始\n9. タイマーが終了するまで待つ\n10. NetworkタブでPOST /api/sessionsとPUT /api/sessions/{id}の呼び出しを確認\n11. データベースでセッションが3件記録されていることを確認（オプション）',
            'expected_screen': '各セッションが正常に開始・終了する\n終了モーダルが毎回表示される\nエラーが発生しない',
            'expected_api': 'POST /api/sessionsが3回呼ばれる\nPUT /api/sessions/{id}が3回呼ばれる\n各セッションに異なるIDが割り当てられる'
        },
        {
            'id': 'ST-005',
            'scenario': 'SCENARIO-004',
            'name': '作業→休憩→作業の連続実行フロー',
            'purpose': 'ポモドーロテクニックの基本フロー（作業→休憩→作業）が正しく動作することを確認',
            'steps': '1. ブラウザでhttp://localhost:5173にアクセス\n2. セッションタイプが「作業時間」と表示されることを確認\n3. 「開始」ボタンをクリック\n4. タイマーが作業時間（1分）からカウントダウンすることを確認\n5. タイマーが0:00になるまで待つ\n6. 終了モーダルが表示されることを確認\n7. セッションタイプが自動的に「休憩時間」に切り替わることを確認\n8. 休憩タイマーが自動的に開始されることを確認（実装されている場合）\n9. 休憩タイマーが0:00になるまで待つ\n10. セッションタイプが自動的に「作業時間」に切り替わることを確認\n11. 次の作業タイマーを開始\n12. 作業タイマーが開始されることを確認',
            'expected_screen': '作業タイマーが正常に実行される\n作業タイマー終了後、自動的に休憩タイマーに切り替わる\n休憩タイマーが正常に実行される\n休憩タイマー終了後、自動的に作業タイマーに切り替わる\nセッションタイプの表示が正しく切り替わる',
            'expected_api': '作業セッション: POST /api/sessions（sessionType: "WORK"）\n作業セッション完了: PUT /api/sessions/{id}（completed: true）\n休憩セッション: POST /api/sessions（sessionType: "BREAK"）\n休憩セッション完了: PUT /api/sessions/{id}（completed: true）'
        },
        {
            'id': 'ST-006',
            'scenario': 'SCENARIO-001,003',
            'name': 'ブラウザリロード時の状態復元',
            'purpose': 'ブラウザをリロードした際に、設定や認証状態が正しく復元されることを確認',
            'steps': '1. ブラウザでhttp://localhost:5173にアクセス\n2. 設定画面で作業時間を30分、休憩時間を10分に変更して保存\n3. Spotify認証を完了（オプション）\n4. タイマーページに戻る\n5. ブラウザをリロード（F5）\n6. 設定値が保持されていることを確認\n7. Spotify認証状態が保持されていることを確認（認証済みの場合）\n8. タイマーの初期表示が変更した設定値になっていることを確認\n9. ApplicationタブでLocalStorageの内容を確認',
            'expected_screen': 'ページリロード後も設定値が保持されている\nSpotify認証状態が保持されている（認証済みの場合）\nタイマーの初期表示が正しい設定値になっている\nエラーが表示されない',
            'expected_api': 'ページリロード時、GET /api/settings/{deviceId}が呼ばれる\n設定情報が正しく取得される'
        },
        {
            'id': 'ST-007',
            'scenario': 'SCENARIO-005',
            'name': 'エラー発生時の動作確認',
            'purpose': '各種エラーが発生した際に、システムが適切にエラーハンドリングを行い、ユーザーに分かりやすいメッセージを表示することを確認',
            'steps': '【ケース1: バリデーションエラー】\n1. 設定画面で作業時間を100分（上限超過）に設定\n2. 「保存」ボタンをクリック\n3. エラーメッセージが表示されることを確認\n\n【ケース2: 認証エラー】\n4. LocalStorageからSpotify認証情報を削除\n5. 「プレイリスト選択」ボタンをクリック\n6. 認証要求またはエラーメッセージが表示されることを確認\n\n【ケース3: ネットワークエラー】\n7. バックエンドサーバーを停止\n8. タイマーページで「開始」ボタンをクリック\n9. エラーメッセージが表示されることを確認\n10. バックエンドサーバーを再起動\n11. 再度「開始」ボタンをクリック\n12. 正常に動作することを確認',
            'expected_screen': 'バリデーションエラー時、具体的なエラーメッセージが表示される\n認証エラー時、認証要求またはエラーメッセージが表示される\nネットワークエラー時、エラーメッセージが表示される\nエラー発生後もアプリがクラッシュしない\nエラー解決後、正常に動作する',
            'expected_api': 'バリデーションエラー: PUT /api/settingsが400 Bad Requestを返す\n認証エラー: GET /api/spotify/playlistsが401/500を返す\nネットワークエラー: API呼び出しが失敗する'
        },
    ]

    create_scenario_test_sheet(ws_scenario_tests, 'シナリオテスト詳細', scenario_tests, header_fill, header_font, border, center_align, left_align)

    # ===== シート4: ユーザビリティテスト =====
    usability_tests = [
        {
            'id': 'ST-008',
            'name': 'モバイル対応確認',
            'purpose': 'スマートフォンやタブレットでの表示・操作性を確認',
            'steps': '1. ブラウザの開発者ツール（F12）を開く\n2. デバイスツールバーを有効化（Ctrl+Shift+M / Cmd+Shift+M）\n3. スマートフォンサイズ（375px幅）に設定\n4. http://localhost:5173にアクセス\n5. タイマーページが正しく表示されることを確認\n6. ボタンがタップしやすいサイズであることを確認\n7. テキストが読みやすいことを確認\n8. 設定画面にアクセス\n9. 設定画面が正しく表示されることを確認\n10. 入力欄が操作しやすいことを確認',
            'expected': 'モバイルサイズでもレイアウトが崩れない\nボタンがタップしやすいサイズである（推奨: 44px以上）\nテキストが読みやすいサイズである\n横スクロールが発生しない\nすべての機能が操作可能である'
        },
        {
            'id': 'ST-009',
            'name': '画面遷移の確認',
            'purpose': '各画面間の遷移がスムーズで直感的であることを確認',
            'steps': '1. タイマーページ（/）から設定画面（/settings）に遷移\n2. 設定画面からタイマーページに戻る（「戻る」リンク）\n3. 設定画面で「保存」ボタンをクリックしてタイマーページに遷移\n4. 設定画面で「キャンセル」ボタンをクリックしてタイマーページに遷移\n5. 各遷移がスムーズであることを確認\n6. URLが正しく変更されることを確認\n7. ブラウザの戻るボタンで前の画面に戻れることを確認',
            'expected': '画面遷移がスムーズである（遅延がない）\nURLが正しく変更される\nブラウザの戻るボタンが動作する\n遷移時にエラーが発生しない'
        },
    ]

    create_usability_sheet(ws_usability, 'ユーザビリティテスト', usability_tests, header_fill, header_font, border, center_align, left_align)

    # ===== シート5: システム全体動作確認 =====
    system_tests = [
        {
            'id': 'ST-010',
            'name': 'データ整合性の確認（複数操作後）',
            'purpose': '複数の操作を実行した後、フロントエンド、バックエンド、データベース間でデータが整合していることを確認',
            'steps': '1. 設定画面で作業時間を30分、休憩時間を10分に変更して保存\n2. タイマーを3回実行（作業→休憩→作業）\n3. 設定を再度変更（作業時間を25分、休憩時間を5分に戻す）して保存\n4. タイマーページで設定が反映されていることを確認\n5. ページをリロード\n6. 設定が保持されていることを確認\n7. データベースで設定テーブルとセッションテーブルの内容を確認（オプション）\n8. フロントエンドの表示とデータベースの内容が一致していることを確認',
            'expected_screen': '設定変更が正しく反映される\nセッションが正しく記録される\nページリロード後も設定が保持される\nデータが失われていない',
            'expected_db': '設定テーブルに最新の設定値が保存されている\nセッションテーブルに3件のセッションが記録されている\n各セッションのcompletedフラグがtrueである\n各セッションのendedAtが設定されている'
        },
    ]

    create_system_sheet(ws_system, 'システム全体動作確認', system_tests, header_fill, header_font, border, center_align, left_align)

    # ===== シート6: テスト結果記録 =====
    ws_results.merge_cells('A1:F1')
    ws_results['A1'] = 'テスト実行記録表'
    ws_results['A1'].font = title_font
    ws_results['A1'].alignment = center_align

    result_headers = ['テスト項目ID', '実行日時', '実行者', '結果', '不具合ID', '備考']
    ws_results.append([])
    ws_results.append(result_headers)

    for col in range(1, len(result_headers) + 1):
        cell = ws_results.cell(row=3, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    # テスト項目IDを追加
    for item in test_items:
        row = [item[1], '', '', '', '', '']
        ws_results.append(row)
        for col in range(1, len(row) + 1):
            cell = ws_results.cell(row=ws_results.max_row, column=col)
            cell.border = border
            if col == 4:  # 結果列
                cell.alignment = center_align
            else:
                cell.alignment = left_align

    # 列幅調整
    ws_results.column_dimensions['A'].width = 15
    ws_results.column_dimensions['B'].width = 20
    ws_results.column_dimensions['C'].width = 15
    ws_results.column_dimensions['D'].width = 10
    ws_results.column_dimensions['E'].width = 15
    ws_results.column_dimensions['F'].width = 40

    # テストサマリー
    ws_results.append([])
    ws_results.append(['テストサマリー'])
    ws_results.cell(row=ws_results.max_row, column=1).font = title_font

    summary_headers = ['項目', '件数']
    ws_results.append(summary_headers)
    for col in range(1, len(summary_headers) + 1):
        cell = ws_results.cell(row=ws_results.max_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    summary_data = [
        ['総テスト項目数', len(test_items)],
        ['合格件数', ''],
        ['不合格件数', ''],
        ['未実施件数', ''],
        ['合格率', ''],
    ]

    for row_data in summary_data:
        ws_results.append(row_data)
        for col in range(1, len(row_data) + 1):
            cell = ws_results.cell(row=ws_results.max_row, column=col)
            cell.border = border
            cell.alignment = left_align if col == 1 else center_align

    # ファイル保存
    filename = '総合テスト項目書_学習用ポモドーロアプリ.xlsx'
    wb.save(filename)
    print(f'Excelファイルを作成しました: {filename}')
    return filename

def create_scenario_test_sheet(ws, title, tests, header_fill, header_font, border, center_align, left_align):
    """シナリオテスト詳細シートを作成"""
    ws.merge_cells('A1:G1')
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center_align

    headers = ['テスト項目ID', 'シナリオID', 'テスト項目名', 'テスト目的', 'テスト手順', '期待結果（画面）', '期待結果（API）']
    ws.append([])
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    for test in tests:
        row = [
            test['id'],
            test['scenario'],
            test['name'],
            test['purpose'],
            test['steps'],
            test['expected_screen'],
            test['expected_api']
        ]
        ws.append(row)
        for col in range(1, len(row) + 1):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.border = border
            cell.alignment = left_align

    # 列幅調整
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 50
    ws.column_dimensions['G'].width = 50

def create_usability_sheet(ws, title, tests, header_fill, header_font, border, center_align, left_align):
    """ユーザビリティテストシートを作成"""
    ws.merge_cells('A1:E1')
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center_align

    headers = ['テスト項目ID', 'テスト項目名', 'テスト目的', 'テスト手順', '期待結果']
    ws.append([])
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    for test in tests:
        row = [
            test['id'],
            test['name'],
            test['purpose'],
            test['steps'],
            test['expected']
        ]
        ws.append(row)
        for col in range(1, len(row) + 1):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.border = border
            cell.alignment = left_align

    # 列幅調整
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 50

def create_system_sheet(ws, title, tests, header_fill, header_font, border, center_align, left_align):
    """システム全体動作確認シートを作成"""
    ws.merge_cells('A1:F1')
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center_align

    headers = ['テスト項目ID', 'テスト項目名', 'テスト目的', 'テスト手順', '期待結果（画面）', '期待結果（データベース）']
    ws.append([])
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    for test in tests:
        row = [
            test['id'],
            test['name'],
            test['purpose'],
            test['steps'],
            test['expected_screen'],
            test['expected_db']
        ]
        ws.append(row)
        for col in range(1, len(row) + 1):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.border = border
            cell.alignment = left_align

    # 列幅調整
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 50

if __name__ == '__main__':
    try:
        create_system_test_excel()
    except ImportError:
        print('エラー: openpyxlライブラリが必要です。')
        print('インストール方法: pip install openpyxl')
    except Exception as e:
        print(f'エラーが発生しました: {e}')


